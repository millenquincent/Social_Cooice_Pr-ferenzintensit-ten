import pandas as pd
import openpyxl, time, numpy as np


Länderjahr        = ['IND_2019', 'USA_2020', 'CHE_2019', 'CAN_2008', 'CHE_2011', 'CAN_2015', 'NZL_1996', 'AUS_2013', 'SWE_2018', 'USA_2016', 'NLD_2021', 'CAN_2011', 'CHE_2007', 'DEU_2021', 'BRA_2014', 'GBR_1997', 'CAN_2019', 'PRT_2005', 'GBR_2019', 'BRA_2002', 'BRA_2018', 'POL_2005', 'MEX_2012', 'MEX_2009', 'IRL_2002', 'NLD_2006', 'ROU_2012', 'BEL_2003', 'BELF1999', 'NLD_2010', 'USA_2008', 'NLD_1998', 'DEU_2009', 'ISL_2017', 'NOR_1997', 'NOR_2001', 'CHE_1999', 'MEX_1997', 'DEU_2017', 'PER_2006', 'SVN_1996', 'DNK_2001', 'TWN_2001', 'DEU_1998', 'DEU_2005', 'FRA_2012', 'NOR_2005', 'POL_1997', 'POL_2019', 'CZE_2006', 'DNK_1998', 'ITA_2018', 'FRA_2007', 'KGZ_2005', 'AUS_2019', 'CHL_2017', 'DEU12002', 'BRA_2010', 'MEX_2003', 'THA_2007', 'JPN_2004', 'BELW1999', 'JPN_2013', 'USA_2012', 'POL_2011', 'ROU_2004', 'TWN_2008', 'NLD_2017', 'DEU_2013', 'AUS_2007', 'CZE_2010', 'IRL_2011', 'CAN_1997', 'RUS_1999', 'FRA_2017', 'TWN_2012', 'TWN_2004', 'POL_2007', 'NZL_2017', 'AUS_1996', 'POL_2001', 'NOR_2017', 'NOR_2009', 'LTU_2020', 'AUS_2004', 'MEX_2000', 'RUS_2000', 'NZL_2002', 'NOR_2013', 'NZL_2020', 'TWN_2016', 'JPN_2017', 'TWN_2020', 'CAN_2004', 'CZE_2013', 'ISL_1999', 'FIN_2019', 'ISL_2007', 'MEX_2006', 'FIN_2015', 'NLD_2002', 'PER_2016', 'PER_2011', 'SRB_2012', 'GBR_2015', 'CZE_2017', 'SWE_2006', 'THA_2019', 'USA_1996', 'HUN_1998', 'PRT_2019', 'KOR_2004', 'THA_2011', 'LTU_2016', 'PRT_2015', 'RUS_2004', 'CZE_2021', 'SLV_2019', 'BGR_2001', 'ISL_2013', 'TUN_2019', 'CRI_2018', 'ISL_2003', 'DNK_2007', 'ITA_2006', 'IRL_2007', 'CHE_2003', 'DNK_2019', 'ARG_2015', 'ROU_2009', 'ISL_2009', 'NZL_2011', 'JPN_2007', 'JPN_1996', 'GRC_2019', 'PRT_2009', 'PRT_2002', 'ZAF_2014', 'FIN_2011', 'ISL_2016', 'FIN_2007', 'ALB_2017', 'MEX_2018', 'CZE_1996', 'NZL_2014', 'MNE_2016', 'ISR_2003', 'ESP_1996', 'ESP_2004', 'ISR_2020', 'HUN_2018', 'ESP_2000', 'ESP_2008', 'SVK_2010', 'AUT_2017', 'HUN_2002', 'ZAF_2009', 'CHL_2005', 'PHL_2010', 'KEN_2013', 'ISR_2006', 'PHL_2004', 'PHL_2016', 'CHL_2009', 'TWN_1996', 'URY_2019', 'KOR_2016', 'PER_2021', 'MEX_2015', 'FIN_2003', 'ROU_1996', 'CHL_1999', 'AUT_2008', 'SWE_1998', 'SVK_2016', 'NZL_2008', 'UKR_1998', 'PER_2001', 'ALB_2005', 'ROU_2014', 'TUR_2011', 'ROU_2016', 'PER_2000', 'KOR_2000', 'ISR_1996', 'TUR_2015', 'BELF2019', 'THA_2001', 'GRC22015', 'TUR_2018', 'USA_2004', 'SWE_2002', 'SVN_2008', 'HKG_2012', 'LVA_2014', 'SVN_2011', 'GRC_2012', 'DEU22002', 'GRC_2009', 'HKG_2016', 'ISR_2013', 'LVA_2018', 'LTU_1997', 'GRC12015', 'LVA_2010', 'HRV_2007', 'LVA_2011', 'SVK_2020', 'SVN_2004', 'EST_2011', 'HKG_1998', 'KOR_2012', 'BLR_2001', 'BRA_2006', 'KOR_2008', 'BLR_2008', 'FRA_2002', 'IRL_2016', 'AUT_2013', 'BGR_2014', 'GBR_2017', 'URY_2009', 'MNE_2012', 'CZE_2002', 'GBR_2005', 'SWE_2014', 'HKG_2008', 'BELW2019', 'HKG_2000', 'HKG_2004']
Länderjahr_Anzahl = [['IND_2019', 13963], ['USA_2020', 7449], ['CHE_2019', 4645], ['CAN_2008', 4495], ['CHE_2011', 4391], ['CAN_2015', 4202], ['NZL_1996', 4080], ['AUS_2013', 3953], ['SWE_2018', 3784], ['USA_2016', 3648], ['NLD_2021', 3485], ['CAN_2011', 3458], ['CHE_2007', 3164], ['DEU_2021', 3152], ['BRA_2014', 3136], ['GBR_1997', 2897], ['CAN_2019', 2889], ['PRT_2005', 2801], ['GBR_2019', 2537], ['BRA_2002', 2514], ['BRA_2018', 2506], ['POL_2005', 2402], ['MEX_2012', 2400], ['MEX_2009', 2400], ['IRL_2002', 2367], ['NLD_2006', 2359], ['ROU_2012', 2283], ['BEL_2003', 2225], ['BELF1999', 2179], ['NLD_2010', 2153], ['USA_2008', 2102], ['NLD_1998', 2101], ['DEU_2009', 2095], ['ISL_2017', 2073], ['NOR_1997', 2055], ['NOR_2001', 2052], ['CHE_1999', 2048], ['MEX_1997', 2033], ['DEU_2017', 2032], ['PER_2006', 2032], ['SVN_1996', 2031], ['DNK_2001', 2026], ['TWN_2001', 2022], ['DEU_1998', 2019], ['DEU_2005', 2018], ['FRA_2012', 2014], ['NOR_2005', 2012], ['POL_1997', 2003], ['POL_2019', 2003], ['CZE_2006', 2002], ['DNK_1998', 2001], ['ITA_2018', 2001], ['FRA_2007', 2000], ['KGZ_2005', 2000], ['AUS_2019', 2000], ['CHL_2017', 2000], ['DEU12002', 2000], ['BRA_2010', 2000], ['MEX_2003', 1991], ['THA_2007', 1990], ['JPN_2004', 1977], ['BELW1999', 1960], ['JPN_2013', 1937], ['USA_2012', 1929], ['POL_2011', 1919], ['ROU_2004', 1913], ['TWN_2008', 1905], ['NLD_2017', 1903], ['DEU_2013', 1889], ['AUS_2007', 1873], ['CZE_2010', 1857], ['IRL_2011', 1853], ['CAN_1997', 1851], ['RUS_1999', 1842], ['FRA_2017', 1830], ['TWN_2012', 1826], ['TWN_2004', 1823], ['POL_2007', 1817], ['NZL_2017', 1808], ['AUS_1996', 1798], ['POL_2001', 1794], ['NOR_2017', 1792], ['NOR_2009', 1782], ['LTU_2020', 1781], ['AUS_2004', 1769], ['MEX_2000', 1766], ['RUS_2000', 1748], ['NZL_2002', 1741], ['NOR_2013', 1727], ['NZL_2020', 1725], ['TWN_2016', 1690], ['JPN_2017', 1688], ['TWN_2020', 1680], ['CAN_2004', 1674], ['CZE_2013', 1653], ['ISL_1999', 1631], ['FIN_2019', 1598], ['ISL_2007', 1595], ['MEX_2006', 1591], ['FIN_2015', 1587], ['NLD_2002', 1574], ['PER_2016', 1572], ['PER_2011', 1570], ['SRB_2012', 1568], ['GBR_2015', 1567], ['CZE_2017', 1559], ['SWE_2006', 1547], ['THA_2019', 1536], ['USA_1996', 1534], ['HUN_1998', 1525], ['PRT_2019', 1500], ['KOR_2004', 1500], ['THA_2011', 1500], ['LTU_2016', 1500], ['PRT_2015', 1499], ['RUS_2004', 1496], ['CZE_2021', 1490], ['SLV_2019', 1488], ['BGR_2001', 1482], ['ISL_2013', 1479], ['TUN_2019', 1477], ['CRI_2018', 1456], ['ISL_2003', 1446], ['DNK_2007', 1442], ['ITA_2006', 1439], ['IRL_2007', 1435], ['CHE_2003', 1418], ['DNK_2019', 1418], ['ARG_2015', 1406], ['ROU_2009', 1403], ['ISL_2009', 1385], ['NZL_2011', 1374], ['JPN_2007', 1373], ['JPN_1996', 1327], ['GRC_2019', 1324], ['PRT_2009', 1316], ['PRT_2002', 1303], ['ZAF_2014', 1300], ['FIN_2011', 1298], ['ISL_2016', 1295], ['FIN_2007', 1283], ['ALB_2017', 1255], ['MEX_2018', 1239], ['CZE_1996', 1229], ['NZL_2014', 1224], ['MNE_2016', 1213], ['ISR_2003', 1212], ['ESP_1996', 1212], ['ESP_2004', 1212], ['ISR_2020', 1209], ['HUN_2018', 1208], ['ESP_2000', 1208], ['ESP_2008', 1204], ['SVK_2010', 1203], ['AUT_2017', 1203], ['HUN_2002', 1200], ['ZAF_2009', 1200], ['CHL_2005', 1200], ['PHL_2010', 1200], ['KEN_2013', 1200], ['ISR_2006', 1200], ['PHL_2004', 1200], ['PHL_2016', 1200], ['CHL_2009', 1200], ['TWN_1996', 1200], ['URY_2019', 1200], ['KOR_2016', 1199], ['PER_2021', 1199], ['MEX_2015', 1197], ['FIN_2003', 1196], ['ROU_1996', 1175], ['CHL_1999', 1173], ['AUT_2008', 1165], ['SWE_1998', 1157], ['SVK_2016', 1150], ['NZL_2008', 1149], ['UKR_1998', 1148], ['PER_2001', 1118], ['ALB_2005', 1116], ['ROU_2014', 1112], ['TUR_2011', 1109], ['ROU_2016', 1105], ['PER_2000', 1102], ['KOR_2000', 1100], ['ISR_1996', 1091], ['TUR_2015', 1086], ['BELF2019', 1084], ['THA_2001', 1081], ['GRC22015', 1078], ['TUR_2018', 1069], ['USA_2004', 1066], ['SWE_2002', 1060], ['SVN_2008', 1055], ['HKG_2012', 1044], ['LVA_2014', 1036], ['SVN_2011', 1031], ['GRC_2012', 1029], ['DEU22002', 1023], ['GRC_2009', 1022], ['HKG_2016', 1020], ['ISR_2013', 1017], ['LVA_2018', 1011], ['LTU_1997', 1009], ['GRC12015', 1008], ['LVA_2010', 1005], ['HRV_2007', 1004], ['LVA_2011', 1004], ['SVK_2020', 1003], ['SVN_2004', 1002], ['EST_2011', 1000], ['HKG_1998', 1000], ['KOR_2012', 1000], ['BLR_2001', 1000], ['BRA_2006', 1000], ['KOR_2008', 1000], ['BLR_2008', 1000], ['FRA_2002', 1000], ['IRL_2016', 1000], ['AUT_2013', 1000], ['BGR_2014', 999], ['GBR_2017', 984], ['URY_2009', 968], ['MNE_2012', 967], ['CZE_2002', 948], ['GBR_2005', 860], ['SWE_2014', 832], ['HKG_2008', 815], ['BELW2019', 730], ['HKG_2000', 674], ['HKG_2004', 582]]
Tabelle_Pandas    = pd.read_excel('14_cses_imd.xlsx')


#Hilfsfunktionen___________________________________________________________________________________
# Einfach Skala Liste entnehmen
def Skala_Reihe_bekommen(Tabelle: pd.core.frame.DataFrame, Reihennummer: int):
    Reihe_vollständig = list(Tabelle.iloc[Reihennummer])
    Reihe_Skala = Reihe_vollständig[1:10]
    return Reihe_Skala
#print(Skala_Reihe_bekommen(Tabelle_Pandas,9))

# Aus Skala Ranking machen [2, 4, 98, 98, 5, 2, 98, 98] als input zu [4, 3, 1, 1, 2, 4, 3, 4]
def Aus_Skala_Ranking_machen(Liste: list):
    unique_sorted_values = sorted(set(Liste), reverse=True)


    New_List = [0]*len(Liste)
    for i in range(0,len(unique_sorted_values)):
        second_highest = unique_sorted_values[i] if len(unique_sorted_values) > 1 else unique_sorted_values[0]
        indices_second_highest = [index for index, value in enumerate(Liste) if value == second_highest]
        for t in range(0,len(indices_second_highest)):
            New_List[indices_second_highest[t]] = i

    count = 0
    for i in range(0,len(Liste)):
        if Liste[i] > 10:
            New_List[i] = 98
            count = count+1
    if count == 0:
        New_List = [x + 1 for x in New_List]

    return New_List
#print(Aus_Skala_Ranking_machen(Reihe_Skala))

#Teil_1_irrelevanten_Zeilen_entfernen_und_Kopfzeile_erstellen______________________________________
# Erste Zeile mit AB(1,2), AB(3,1) unds so hinzufügen in Excel
Tabelle_Pandas        = '0_cses_imd.xlsx'
def Um_Zeile_null_zu_schreiben(Excelname):
    Tabelle_Openpyxl = openpyxl.load_workbook(Excelname)
    Sheet = Tabelle_Openpyxl['Sheet1']
    Kopfzeile_1_bis19 = ['Country','SA','SB','SC','SD','SE','SF','SG','SH','SI','A','B','C','D','E','F','G','H','I']

    for i in range(0,len(Kopfzeile_1_bis19)):
        Sheet.cell(row=1, column=i+1, value= Kopfzeile_1_bis19[i])

    Liste = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    Neue_Liste = []
    for i in Liste:
        for t in Liste:
            if t != i:
                Neue_Liste.append(i+t)


    Liste_Werte_Index = []
    for i in range(0,len(Neue_Liste)):
        Liste_Werte_Index.append([(i*2)+20,Neue_Liste[i]+'(1,2)'])
    #print(Liste_Werte_Index)
    for i in Liste_Werte_Index:
        Sheet.cell(row=1, column=i[0], value= i[1])

    Liste_zwischen_werte_Index = []
    for i in range(0,len(Neue_Liste)):
        Liste_zwischen_werte_Index.append([(i*2)+21,Neue_Liste[i]+'(3,1)'])
    #print(Liste_zwischen_werte_Index)
    for i in Liste_zwischen_werte_Index:
        Sheet.cell(row=1, column=i[0], value= i[1])

    Tabelle_Openpyxl.save(Excelname)
    Tabelle_Openpyxl.close()
    return 'KopfZeile hinzugefügt!'
#print(Um_Zeile_null_zu_schreiben(Tabelle_Pandas))

Tabelle_Pandas        = pd.read_excel('10_cses_imd.xlsx')
# Reihen wo bei Skala nur zwei Werte sind rauslöschen
def Irrelevanten_Reihen_löschen1(Tabelle_Pandas):
    zu_löschende_index = []
    for i in range(0,len(Tabelle_Pandas)):
        Reihenliste = list(Tabelle_Pandas.iloc[i,1:10])
        if sorted(Reihenliste)[2] > 10: zu_löschende_index.append(i)

    Tabelle_Pandas = Tabelle_Pandas.drop(zu_löschende_index)
    Tabelle_Pandas.to_excel('11_cses_imd.xlsx', index=False)
    return 'Irrelevanten Reihen gelöscht!'
#print(Irrelevanten_Reihen_löschen1(Tabelle_Pandas))

Tabelle_Pandas        = pd.read_excel('11_cses_imd.xlsx')
# Reihen wo bei Skala nur 3 werte hat wovon einer 98 ist
def Irrelevanten_Reihen_löschen2(Tabelle_Pandas):
    zu_löschende_index = []
    for i in range(0,len(Tabelle_Pandas)):
        Reihenliste = list(set(Tabelle_Pandas.iloc[i,1:10]))
        sorted_list = sorted(Reihenliste)
        lenght = len(sorted_list)
        if lenght < 3 or lenght ==3 and sorted_list[-1] > 10: zu_löschende_index.append(i)
            #print([Reihenliste,i])


    Tabelle_Pandas = Tabelle_Pandas.drop(zu_löschende_index)
    Tabelle_Pandas.to_excel('12_cses_imd.xlsx', index=False)
    return 'Irrelevanten Reihen gelöscht!'
#print(Irrelevanten_Reihen_löschen2(Tabelle_Pandas))


#Teil_2_Rankings_und_Differenzen_hinzufügen________________________________________________________
Tabelle_Pandas        = pd.read_excel('12_cses_imd.xlsx')
def Rankings_Pandas_Tabelle_hinzufügen(Tabelle_Pandas):

    for i in range(0,len(Tabelle_Pandas)):
        Reihe_Skala   = Skala_Reihe_bekommen(Tabelle_Pandas,i)
        Reihe_Ranking = Aus_Skala_Ranking_machen(Reihe_Skala)

        for t in range(0,9):
            Tabelle_Pandas.iloc[i, t + 10] = Reihe_Ranking[t] #Reihe Spalte
    Tabelle_Pandas.to_excel('13_cses_imd.xlsx', index=False)
    return '1) Rankings hinzugefügt!'
#print(Rankings_Pandas_Tabelle_hinzufügen(Tabelle_Pandas))

Tabelle_Pandas        = pd.read_excel('13_cses_imd.xlsx')
def Ranking_zu_Skalen_finden_wo_XY_passt(Tabelle_Pandas):

    for reihe in range(0,len(Tabelle_Pandas)):
        Spaltewert1 = 18
        Spaltewert2 = 19
        for i in range(10,19):
            for t in range(10,19):
                if i != t:

                    Rankingwert1 = Tabelle_Pandas.iloc[reihe,i]
                    Rankingwert2 = Tabelle_Pandas.iloc[reihe,t]
                    Skalawert1 = Tabelle_Pandas.iloc[reihe,i-9]
                    Skalawert2 = Tabelle_Pandas.iloc[reihe,t-9]
                    Spaltewert1 = Spaltewert1 +2
                    Spaltewert2 = Spaltewert2 +2
                    #print(f'Ranking: {[Rankingwert1,Rankingwert2]}')
                    #print(f'Skala: {[Skalawert1,Skalawert2]}')
                    #print(f'Spalte: {[Spaltewert1,Spaltewert2]}')

                    if Rankingwert1 == 1 and Rankingwert2 == 2:
                        Skaladifferenz = abs(Skalawert1 - Skalawert2)
                        Tabelle_Pandas.iloc[reihe, Spaltewert1-1] = Skaladifferenz
                        #print([reihe, Spaltewert1])

                    if Rankingwert1 == 3 and Rankingwert2 == 1:
                        Skaladifferenz = abs(Skalawert1 - Skalawert2)
                        Tabelle_Pandas.iloc[reihe, Spaltewert2-1] = Skaladifferenz
                        #print([reihe, Spaltewert2])

                    #print(reihe)
    Tabelle_Pandas.to_excel('14_cses_imd.xlsx', index=False)
    return '2) Differenzen hinzugefügt!'
#print(Ranking_zu_Skalen_finden_wo_XY_passt(Tabelle_Pandas))


Tabelle_Pandas        = pd.read_excel('14_cses_imd.xlsx')
#Teil_3_Tabellen_für_alle_Länderjahre_erstellen_(1,2)_und_(3,1)____________________________________
def Summen_und_Anzahl_Durchschnitt_bekommen(Tabelle_Pandas,Landname):

    #Für XY (1,2)
    Kopfzeile = list(Tabelle_Pandas.columns)
    Kopfzeile[0] = 'Präferenz_Kombination_der_Rankings_davon_Differenz_der_Skalenwerte'
    Erste_19_der_Kopfzeile = Kopfzeile[:19]
    for i in range(19,len(Kopfzeile),2):
        Erste_19_der_Kopfzeile.append(Kopfzeile[i])
    Kopfzeile = Erste_19_der_Kopfzeile

    Summe_li               = ['Summe',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan]
    for i in range(19,Tabelle_Pandas.shape[1],2):
        Spalten_Summe = Tabelle_Pandas.iloc[:,i].sum()
        Summe_li.append(Spalten_Summe)

    Anzahl_Werte_li = ['Absolute_Anzahl',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan]
    for i in range(19,Tabelle_Pandas.shape[1],2):
        Spalten_Summe = Tabelle_Pandas.iloc[:,i].count()
        Anzahl_Werte_li.append(Spalten_Summe)

    
    Durchschnitt_li = ['Durchschnitt',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan]
    for i in range(19,len(Anzahl_Werte_li)):
        if Anzahl_Werte_li[i] == 0:
            Durchschnitt_li.append(0)
        else:
            Durchschnitt_Spalte = Summe_li[i]/Anzahl_Werte_li[i]
            Durchschnitt_li.append(Durchschnitt_Spalte)

    #Neues Dataframe erstellen
    Ergebnisse = {
        'Kopfzeile'      : Kopfzeile,
        'Summen'         : Summe_li,
        'Absolute Anzahl': Anzahl_Werte_li,
        'Durchschnitt'   : Durchschnitt_li}
    Ergebnisse_pandas = pd.DataFrame(Ergebnisse).transpose()
    Excelname = ''+Landname+'Ergebnisse XY(1,2).xlsx'
    Ergebnisse_pandas.to_excel(Excelname, index=False)

    #Für XY (3,1)

    Kopfzeile = list(Tabelle_Pandas.columns)
    Kopfzeile[0] = 'Präferenz_Kombination_der_Rankings_davon_Differenz_der_Skalenwerte'
    Erste_19_der_Kopfzeile = Kopfzeile[:19]
    for i in range(20,len(Kopfzeile),2):
        Erste_19_der_Kopfzeile.append(Kopfzeile[i])
    Kopfzeile = Erste_19_der_Kopfzeile

    Summe_li               = ['Summe',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan]
    for i in range(20,Tabelle_Pandas.shape[1],2):
        Spalten_Summe = Tabelle_Pandas.iloc[:,i].sum()
        Summe_li.append(Spalten_Summe)

    Anzahl_Werte_li = ['Absolute_Anzahl',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan]
    for i in range(20,Tabelle_Pandas.shape[1],2):
        Spalten_Summe = Tabelle_Pandas.iloc[:,i].count()
        Anzahl_Werte_li.append(Spalten_Summe)

    
    Durchschnitt_li = ['Durchschnitt',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan]
    for i in range(19,len(Anzahl_Werte_li)):
        if Anzahl_Werte_li[i] == 0:
            Durchschnitt_li.append(0)
        else:
            Durchschnitt_Spalte = Summe_li[i]/Anzahl_Werte_li[i]
            Durchschnitt_li.append(Durchschnitt_Spalte)

    #Neues Dataframe erstellen
    Ergebnisse = {
        'Kopfzeile'      : Kopfzeile,
        'Summen'         : Summe_li,
        'Absolute Anzahl': Anzahl_Werte_li,
        'Durchschnitt'   : Durchschnitt_li}
    Ergebnisse_pandas = pd.DataFrame(Ergebnisse).transpose()
    Excelname = ''+Landname+'Ergebnisse XY(3,1).xlsx'
    Ergebnisse_pandas.to_excel(Excelname, index=False)

    return Landname+'Summe, Anzahl und Durchschnitt in Excel für XY (1,2) und XY (3,1) hinzugefügt!'
#print(Summen_und_Anzahl_Durchschnitt_bekommen(Tabelle_Pandas,'Deutschland'))

Tabelle_Pandas        = pd.read_excel('14_cses_imd.xlsx')
def Excel_mit_reihen_aus_einem_Land(Tabelle_Pandas,Liste_Ländernamen):
    for i in range(0,len(Liste_Ländernamen)):
        Landname = Liste_Ländernamen[i]
        Tabelle_Pandas_Land = Tabelle_Pandas[Tabelle_Pandas['Country'] == Landname]
        Summen_und_Anzahl_Durchschnitt_bekommen(Tabelle_Pandas_Land,Landname)

    return 'für alle Länder zwei Ergebnis Excels erstellt'
#print(Excel_mit_reihen_aus_einem_Land(Tabelle_Pandas,Länderjahr))


#Teil_4_Alle_(1,2)_und_(3,1)_Tabellen_Zusammen_auswerten___________________________________________
def Gesamtauswertung_Summe_Anzahl_Durchschnitt(Länderjahr):

    #Für XY (1,2)
    Zeile1 = ['Summe']+[0]*90
    Zeile2 = ['Absolute_Anzahl']+[0]*90
    Zeile3 = ['Durchschnitt']+[0]*90
    for i in range(len(Länderjahr)):
        Tabelle = pd.read_excel(Länderjahr[i]+'Ergebnisse XY(1,2).xlsx')

        Neue_Zeile1 = Tabelle.iloc[1]
        Zeile1 = Zeile1[1:] + Neue_Zeile1[1:]

        Neue_Zeile2 = Tabelle.iloc[2]
        Zeile2 = Zeile2[1:] + Neue_Zeile2[1:]

    Beispiel     = pd.read_excel('ALB_2005Ergebnisse XY(1,2).xlsx')
    Nullte_Zeile = Beispiel.iloc[0].tolist()
    Erste_Zeile  = pd.concat([pd.Series(['Summe'])          ,Zeile1], ignore_index=True)
    Zweite_Zeile = pd.concat([pd.Series(['Absolute_Anzahl']),Zeile2], ignore_index=True)
    Dritte_Zeile = pd.concat([pd.Series(['Durchschnitt'])   ,Zeile1 / Zeile2], ignore_index=True)


    Auswertung_Tabelle = pd.DataFrame([Nullte_Zeile, Erste_Zeile, Zweite_Zeile, Dritte_Zeile])
    Auswertung_Tabelle.to_excel('00_Gesamtauswertung XY(1,2).xlsx', index=False)

    #Für XY (3,1)
    Zeile1 = ['Summe']+[0]*90
    Zeile2 = ['Absolute_Anzahl']+[0]*90
    Zeile3 = ['Durchschnitt']+[0]*90
    for i in range(len(Länderjahr)):
        Tabelle = pd.read_excel(Länderjahr[i]+'Ergebnisse XY(3,1).xlsx')

        Neue_Zeile1 = Tabelle.iloc[1]
        Zeile1 = Zeile1[1:] + Neue_Zeile1[1:]

        Neue_Zeile2 = Tabelle.iloc[2]
        Zeile2 = Zeile2[1:] + Neue_Zeile2[1:]

    Beispiel     = pd.read_excel('ALB_2005Ergebnisse XY(3,1).xlsx')
    Nullte_Zeile = Beispiel.iloc[0].tolist()
    Erste_Zeile  = pd.concat([pd.Series(['Summe'])          ,Zeile1], ignore_index=True)
    Zweite_Zeile = pd.concat([pd.Series(['Absolute_Anzahl']),Zeile2], ignore_index=True)
    Dritte_Zeile = pd.concat([pd.Series(['Durchschnitt'])   ,Zeile1 / Zeile2], ignore_index=True)


    Auswertung_Tabelle = pd.DataFrame([Nullte_Zeile, Erste_Zeile, Zweite_Zeile, Dritte_Zeile])
    Auswertung_Tabelle.to_excel('00_Gesamtauswertung XY(3,1).xlsx', index=False)

    return 'Auswertungstabellen für (1,2) und (3,1) erstellt'
#print(Gesamtauswertung_Summe_Anzahl_Durchschnitt(Länderjahr))


#__________________________________________________________________________________________________

